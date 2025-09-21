# import streamlit as st
# import openpyxl
# from openpyxl import Workbook
# import os
# import tempfile
# import zipfile
# import pandas as pd

# # ===== Helper: Copy cell with style =====
# def copy_cell(source_cell, target_cell):
#     target_cell.value = source_cell.value
#     if source_cell.has_style:
#         target_cell.font = source_cell.font.copy()
#         target_cell.border = source_cell.border.copy()
#         target_cell.fill = source_cell.fill.copy()
#         target_cell.number_format = source_cell.number_format
#         target_cell.protection = source_cell.protection.copy()
#         target_cell.alignment = source_cell.alignment.copy()

# # ===== Function: Split Excel by column =====
# def split_excel_by_column(ws, headers, column_name, output_dir):
#     if column_name not in headers:
#         st.error(f"❌ Column '{column_name}' not found!")
#         return []

#     col_idx = headers.index(column_name) + 1  # 1-based index

#     # Group rows by column value
#     groups = {}
#     for row in ws.iter_rows(min_row=2):
#         value = row[col_idx - 1].value
#         key = str(value) if value is not None else "Unknown"
#         # Make safe for filename
#         key = "".join(c if c not in '<>:"/\\|?*' else "_" for c in key)
#         if key not in groups:
#             groups[key] = []
#         groups[key].append(row)

#     saved_files = []

#     # Create new file for each group
#     for value, rows in groups.items():
#         new_wb = Workbook()
#         new_ws = new_wb.active
#         new_ws.title = ws.title

#         # Copy headers
#         for i, cell in enumerate(ws[1], start=1):
#             new_cell = new_ws.cell(row=1, column=i)
#             copy_cell(cell, new_cell)
#             if cell.column_letter in ws.column_dimensions:
#                 new_ws.column_dimensions[new_cell.column_letter].width = ws.column_dimensions[cell.column_letter].width

#         # Copy data rows
#         for r_idx, row in enumerate(rows, start=2):
#             for c_idx, cell in enumerate(row, start=1):
#                 new_cell = new_ws.cell(row=r_idx, column=c_idx)
#                 copy_cell(cell, new_cell)

#         # Save file
#         file_name = f"{value}.xlsx"
#         save_path = os.path.join(output_dir, file_name)
#         new_wb.save(save_path)
#         saved_files.append(save_path)

#     return saved_files

# # ===== Create ZIP =====
# def create_zip(files, zip_name):
#     with zipfile.ZipFile(zip_name, 'w') as zf:
#         for f in files:
#             zf.write(f, os.path.basename(f))
#     return zip_name

# # ===== Streamlit UI =====
# st.set_page_config(page_title="Excel Splitter", page_icon="📊", layout="centered")
# st.title("📊 Excel Splitter by Column")
# st.write("Upload an Excel file, choose a column, and split it into multiple files — one per unique value!")

# # File uploader
# uploaded_file = st.file_uploader("📤 Upload your Excel file (.xlsx)", type=["xlsx"])

# if uploaded_file:
#     # Load workbook
#     with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
#         tmp.write(uploaded_file.getvalue())
#         tmp_path = tmp.name

#     wb = openpyxl.load_workbook(tmp_path)
#     ws = wb.active
#     headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

#     st.write("### 📑 Detected Columns:")
#     st.write(", ".join([f"`{h}`" for h in headers]))

#     # Column selector
#     selected_column = st.selectbox("SplitOptions: Split by column ➡️", headers)

#     # Show preview (optional)
#     if st.checkbox("🔍 Show Data Preview (first 5 rows)"):
#         df = pd.read_excel(uploaded_file)
#         st.dataframe(df.head())

#     # Split button
#     if st.button("✂️ Split Excel Now"):
#         with st.spinner("Splitting your Excel file..."):
#             # Create temp dir for output files
#             with tempfile.TemporaryDirectory() as tmpdir:
#                 saved_files = split_excel_by_column(ws, headers, selected_column, tmpdir)

#                 if saved_files:
#                     # Create ZIP
#                     zip_path = os.path.join(tempfile.gettempdir(), "split_excel_files.zip")
#                     create_zip(saved_files, zip_path)

#                     # Provide download
#                     with open(zip_path, "rb") as f:
#                         st.download_button(
#                             label="📥 Download All Split Files (ZIP)",
#                             data=f,
#                             file_name="split_excel_files.zip",
#                             mime="application/zip"
#                         )
#                     st.success(f"✅ Created {len(saved_files)} files!")
#                 else:
#                     st.error("❌ No files created. Check column or data.")

#     # Cleanup temp file
#     os.unlink(tmp_path)

# st.markdown("---")
# st.caption("Built with ❤️ using Streamlit + OpenPyXL")




























import streamlit as st
import openpyxl
from openpyxl import Workbook
import os
import tempfile
import zipfile
import pandas as pd

# ===== CONFIG =====
st.set_page_config(
    page_title="Excel Splitter Pro",
    page_icon="✂️",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# ===== STYLING =====
st.markdown("""
    <style>
    .main-title {
        font-size: 2.5rem;
        font-weight: 700;
        color: #0f52ba;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    .subtitle {
        text-align: center;
        color: #555;
        margin-bottom: 2rem;
    }
    .step-card {
        background: #f9f9ff;
        padding: 1.2rem;
        border-radius: 12px;
        margin: 1rem 0;
        border-left: 4px solid #0f52ba;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    .stButton>button {
        width: 100%;
        background-color: #0f52ba;
        color: white;
        font-weight: 600;
        border-radius: 8px;
        padding: 0.6rem 0;
        border: none;
    }
    .stButton>button:hover {
        background-color: #0a3d8c;
    }
    .stDownloadButton>button {
        width: 100%;
        background-color: #28a745;
        color: white;
        font-weight: 600;
        border-radius: 8px;
        padding: 0.6rem 0;
        border: none;
    }
    .success-box {
        padding: 1rem;
        background-color: #d4edda;
        color: #155724;
        border-radius: 8px;
        border: 1px solid #c3e6cb;
        margin: 1rem 0;
    }
    </style>
""", unsafe_allow_html=True)

# ===== HELPER: Copy cell with style =====
def copy_cell(source_cell, target_cell):
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection.copy()
        target_cell.alignment = source_cell.alignment.copy()

# ===== FUNCTION: Split Excel by column =====
def split_excel_by_column(ws, headers, column_name, output_dir):
    if column_name not in headers:
        return None, f"❌ Column '{column_name}' not found!"

    col_idx = headers.index(column_name) + 1

    groups = {}
    for row in ws.iter_rows(min_row=2):
        value = row[col_idx - 1].value
        key = str(value) if value is not None else "Unknown"
        key = "".join(c if c not in '<>:"/\\|?*' else "_" for c in key)
        if key not in groups:
            groups[key] = []
        groups[key].append(row)

    saved_files = []

    for value, rows in groups.items():
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = ws.title

        # Copy headers
        for i, cell in enumerate(ws[1], start=1):
            new_cell = new_ws.cell(row=1, column=i)
            copy_cell(cell, new_cell)
            if cell.column_letter in ws.column_dimensions:
                new_ws.column_dimensions[new_cell.column_letter].width = ws.column_dimensions[cell.column_letter].width

        # Copy data rows
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, cell in enumerate(row, start=1):
                new_cell = new_ws.cell(row=r_idx, column=c_idx)
                copy_cell(cell, new_cell)

        # Copy merged cells
        for merged_range in ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))

        file_name = f"{value}.xlsx"
        save_path = os.path.join(output_dir, file_name)
        new_wb.save(save_path)
        saved_files.append(save_path)

    return saved_files, None

# ===== FUNCTION: Create ZIP =====
def create_zip(files, zip_name):
    with zipfile.ZipFile(zip_name, 'w') as zf:
        for f in files:
            zf.write(f, os.path.basename(f))
    return zip_name

# ===== APP UI =====
st.markdown('<h1 class="main-title">✂️ Excel Splitter Pro</h1>', unsafe_allow_html=True)
st.markdown('<p class="subtitle">Split your Excel files by any column — beautifully and effortlessly</p>', unsafe_allow_html=True)

# STEP 1: Upload
with st.container():
    st.markdown('<div class="step-card">', unsafe_allow_html=True)
    st.subheader("📤 Step 1: Upload Your Excel File")
    st.info("Supports `.xlsx` files only. All formatting (colors, fonts, widths) will be preserved!")
    uploaded_file = st.file_uploader("", type=["xlsx"], label_visibility="collapsed")
    st.markdown('</div>', unsafe_allow_html=True)

if not uploaded_file:
    st.stop()

# Load workbook
with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
    tmp.write(uploaded_file.getvalue())
    tmp_path = tmp.name

try:
    wb = openpyxl.load_workbook(tmp_path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # STEP 2: Show Columns
    with st.container():
        st.markdown('<div class="step-card">', unsafe_allow_html=True)
        st.subheader("📋 Step 2: Select Column to Split By")
        selected_column = st.selectbox(
            "Choose the column that defines how to split your data:",
            headers,
            help="Each unique value in this column will become a separate Excel file."
        )
        st.markdown('</div>', unsafe_allow_html=True)

    # STEP 3: Preview (Optional)
    with st.expander("🔍 Optional: Preview First 5 Rows", expanded=False):
        df = pd.read_excel(uploaded_file)
        st.dataframe(df.head(), use_container_width=True)

    # STEP 4: Split Button
    if st.button("🚀 Split My Excel Now", use_container_width=True):
        with st.spinner("✂️ Processing your file... This may take a moment for large files."):
            with tempfile.TemporaryDirectory() as tmpdir:
                saved_files, error = split_excel_by_column(ws, headers, selected_column, tmpdir)

                if error:
                    st.error(error)
                elif saved_files:
                    zip_path = os.path.join(tempfile.gettempdir(), "split_excel_files.zip")
                    create_zip(saved_files, zip_path)

                    st.markdown('<div class="success-box">', unsafe_allow_html=True)
                    st.markdown(f"✅ **Success!** Created **{len(saved_files)}** files based on column: `{selected_column}`")
                    st.markdown('</div>', unsafe_allow_html=True)

                    with open(zip_path, "rb") as f:
                        st.download_button(
                            label="📥 Download All Files (ZIP)",
                            data=f,
                            file_name="split_excel_files.zip",
                            mime="application/zip",
                            use_container_width=True
                        )

                    # Show sample filenames
                    with st.expander("📁 Sample Output Filenames", expanded=True):
                        for fname in saved_files[:5]:
                            st.code(os.path.basename(fname))
                        if len(saved_files) > 5:
                            st.caption(f"... and {len(saved_files) - 5} more files")

except Exception as e:
    st.error(f"⚠️ Error processing file: {str(e)}")
    st.info("Please make sure you uploaded a valid .xlsx file.")

finally:
    if 'tmp_path' in locals():
        os.unlink(tmp_path)

# FOOTER
st.markdown("---")
col1, col2, col3 = st.columns([1,2,1])
with col2:
    st.caption("✨ Built with Streamlit + OpenPyXL | Preserves all formatting & styles")
    st.caption("💡 Tip: Avoid special characters in split column for clean filenames")
